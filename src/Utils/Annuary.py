import openpyxl
import openpyxl.utils
import openpyxl.styles
import json

from Utils.Jobs import JobAnvil
from Utils.Reputations import Reputation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle


def init_databases(connection):
    cursor = connection.cursor()
    query = ""
    with open("src/build.sql") as stream:
        for line in stream:
            query += line
    print(query)
    cursor.execute(query, ())
    connection.commit()


def get_json(path):
    with open(path, "rb") as stream:
        return json.load(stream)


def triple_to_rgb(rgb):
    if rgb[0] < 0:
        rgb = (0, rgb[1], rgb[2])
    elif rgb[0] > 255:
        rgb = (255, rgb[1], rgb[2])
    red = format(int(rgb[0]), 'x')
    if len(red) == 1:
        red = "0" + red

    if rgb[1] < 0:
        rgb = (rgb[0], 0, rgb[2])
    elif rgb[1] > 255:
        rgb = (rgb[0], 255, rgb[2])
    green = format(int(rgb[1]), 'x')
    if len(green) == 1:
        green = "0" + green

    if rgb[2] < 0:
        rgb = (rgb[0], rgb[1], 0)
    elif rgb[2] > 255:
        rgb = (rgb[0], rgb[1], 255)
    blue = format(int(rgb[2]), 'x')
    if len(blue) == 1:
        blue = "0" + blue
    return red + green + blue


def store_annuary(path, **kwargs):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if "characters" in kwargs:
        print("Storing characters ...")
        ws_chars = wb.create_sheet(title="Personnages")

        ws_chars[openpyxl.utils.get_column_letter(2) + "1"] = "Nom"
        ws_chars[openpyxl.utils.get_column_letter(3) + "1"] = "Classe"
        ws_chars[openpyxl.utils.get_column_letter(4) + "1"] = "Couleur"
        ws_chars[openpyxl.utils.get_column_letter(5) + "1"] = "Niveau"

        i = 2
        for character in kwargs["characters"]:
            ws_chars[openpyxl.utils.get_column_letter(2) + str(i)] = character.name
            ws_chars[openpyxl.utils.get_column_letter(3) + str(i)] = character.class_
            color_cell = openpyxl.utils.get_column_letter(4) + str(i)
            ws_chars[color_cell] = character.main_trait
            if character.main_trait == "rouge":
                ws_chars[color_cell].font = openpyxl.styles.Font(color=openpyxl.styles.colors.RED, bold=True)
            elif character.main_trait == "bleu":
                ws_chars[color_cell].font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE, bold=True)
            elif character.main_trait == "jaune":
                ws_chars[color_cell].font = openpyxl.styles.Font(color=openpyxl.styles.colors.YELLOW, bold=True)
            ws_chars[openpyxl.utils.get_column_letter(5) + str(i)] = character.level
            i += 1

    if "jobs" in kwargs and "anvils" in kwargs:
        print("Storing jobs & masteries ...")
        ws_jobs = wb.create_sheet(title="Métiers")

        ws_jobs[openpyxl.utils.get_column_letter(2) + "1"] = "Nom"
        ws_jobs[openpyxl.utils.get_column_letter(3) + "1"] = "Métier"
        for idx, mastery in enumerate(JobAnvil.accepted_tiers):
            ws_jobs[openpyxl.utils.get_column_letter(4 + idx) + "1"] = mastery.capitalize()
        anvils_map = {}
        for anvil in kwargs["anvils"]:
            if anvil.id_ not in anvils_map:
                anvils_map[anvil.id_] = []
            anvils_map[anvil.id_].append(anvil)
        i = 2
        for idx, job in enumerate(kwargs["jobs"]):
            ws_jobs[openpyxl.utils.get_column_letter(2) + str(i + idx)] = job.name
            ws_jobs[openpyxl.utils.get_column_letter(3) + str(i + idx)] = job.job
            for ma_idx, mastery in enumerate(JobAnvil.accepted_tiers):
                for anvil in anvils_map[job.id_]:
                    if anvil.tier == mastery:
                        cell = openpyxl.utils.get_column_letter(4 + ma_idx) + str(i + idx)
                        ws_jobs[cell] = mastery.capitalize()
                        if anvil.gold or anvil.bronze:
                            if anvil.gold:
                                ft_color = openpyxl.styles.colors.DARKYELLOW
                                bg_color = openpyxl.styles.colors.YELLOW
                            else:
                                ft_color = openpyxl.styles.colors.DARKRED
                                bg_color = openpyxl.styles.colors.RED
                            ws_jobs[cell].font = openpyxl.styles.Font(color=ft_color, bold=True)
                            ws_jobs[cell].fill = openpyxl.styles.PatternFill(start_color=bg_color, end_color=bg_color,
                                                                             fill_type="solid")
                        break

    if "reputations" in kwargs:
        print("Storing reputations ...")
        ws_reps = wb.create_sheet(title="Réputations")

        for idx, faction in enumerate(Reputation.accepted_factions):
            ws_reps[openpyxl.utils.get_column_letter(3 + idx) + "1"] = faction.capitalize().replace('_', ' ')

        chars_map = {}
        for reputation in kwargs["reputations"]:
            if reputation.name not in chars_map:
                chars_map[reputation.name] = []
            chars_map[reputation.name].append(reputation)

        i = 2
        idx = 0
        for name, reputations in chars_map.items():
            ws_reps[openpyxl.utils.get_column_letter(2) + str(i)] = name.capitalize()
            for idx, faction in enumerate(Reputation.accepted_factions):
                for reputation in reputations:
                    if reputation.faction == faction:
                        ws_reps[openpyxl.utils.get_column_letter(3 + 1 + idx) + str(i)] = reputation.level.capitalize()
                        break
        initial_color = (128, 0, 0)
        final_color = (153, 255, 102)
        n_steps = len(Reputation.accepted_levels)
        step = ()
        for i in range(3):
            step += (int((initial_color[i] + final_color[i]) / n_steps),)
        for idx_level, level in enumerate(Reputation.accepted_levels):
            color = triple_to_rgb( ( (idx_level+1)*step[0], (idx_level+1)*step[1], (idx_level+1)*step[2] ) )
            print(color)
            fill = openpyxl.styles.PatternFill(bgColor=color)
            dxf = DifferentialStyle(fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=level.capitalize(), dxf=dxf)
            ws_reps.conditional_formatting.add("C2:" + openpyxl.utils.get_column_letter(3 + idx) + str(i), rule)

    try:
        wb.save(filename=path)
        print("Register stored !")
    except Exception:
        print("Could not save annuary as xlsx")
